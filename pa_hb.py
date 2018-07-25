#!/usr/bin/env python
# coding=utf-8

import pprint
import csv
# import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import click 
import requests
import datetime as datetime
from datetime import date
from xml.etree import ElementTree as ET
import os
# from random import sample
import random
# import logging
import json
from json import JSONDecodeError

def validate_d(date_text):
	try:
		datetime.datetime.strptime(date_text, '%Y-%m-%d')
	except ValueError:
		raise ValueError("Incorrect data format, should be YYYY-MM-DD")

def daterange(start_date, end_date):
    for n in range(int ((end_date - start_date).days)):
        yield start_date + datetime.timedelta(n)

# booking_id_secret = None
# with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'secrets.json')) as data_file:    
# 	booking_id_secret = (json.load(data_file))['booking_id']

REF_API = 'api'
REF_CLIENT = 'client'
REF_AGENT = 'agent'

TYPE_DEPARTURE = 'departure'
TYPE_CREATION = 'creation'

@click.command()
# @click.option('--days', default=0, type=int)
# @click.option('--duration', default=0, type=int)
@click.option('--hb_pa_file', default='HB二次可订失败20180625.xlsx')
# @click.option('--d_type', default=TYPE_DEPARTURE)
# def pa_hb(days, duration, d_type):
def pa_hb(hb_pa_file):

	# wb2 = load_workbook('test.xlsx')
	try:
		wb = load_workbook(hb_pa_file)
	except FileNotFoundError:
		print('File not found.. Specify file in --hb_pa_file option')
		return

	ws = wb.active
	errors = []
	for i, t in enumerate(tuple(ws.rows)):
		if i == 0:
			continue
		print('i: ' + str(i))
		print(t[0].value)
		if t[0].value == None or t[1].value == None:
			continue

		try:
			t0 = json.loads(t[0].value)
		except JSONDecodeError:
			print('tuple 0.. json load error..')
			continue
		try:
			t1 = json.loads(t[1].value)
		except JSONDecodeError:
			print('tuple 1.. json load error..')
			continue

		ent = {}
		ent['req'] = t0
		ent['res'] = t1
		errors.append(ent)

	err_hotels = []
	err_res = []

	for ent in errors:
		if 'hotel' not in ent['res'].keys():
			continue
		if ent['res']['hotel']['name'] not in err_hotels:
			err_hotels.append(ent['res']['hotel']['name'])
			err_ent = {}
			err_ent['hotel'] = ent['res']['hotel']['name']
			err_ent['code'] = ent['res']['hotel']['code']
			err_ent['count'] = 1
			err_ent['rps'] = set()
			for rp in ent['req']['rooms']:
				err_ent['rps'].add(rp['rateKey'])			
			err_ent['parsed_from'] = hb_pa_file
			err_res.append(err_ent)
		else:
			for err_ent in err_res:
				if ent['res']['hotel']['name'] == err_ent['hotel']:
					err_ent['count'] = err_ent['count'] + 1
					for rp in ent['req']['rooms']:
						err_ent['rps'].add(rp['rateKey'])

	for err_ent in err_res:
		err_ent['rps'] = ', '.join([ val for val in err_ent['rps']])

	output_filename = '_'.join(['output_pa_hb', datetime.datetime.now().strftime('%y%m%d'), 
									datetime.datetime.now().strftime('%H%M'), '.csv'])

	keys = err_res[0].keys()
	with open(output_filename, 'w', newline='', encoding='utf-8') as output_file:
		dict_writer = csv.DictWriter(output_file, keys)
		dict_writer.writeheader()
		dict_writer.writerows(err_res)

if __name__ == '__main__':
	pa_hb()